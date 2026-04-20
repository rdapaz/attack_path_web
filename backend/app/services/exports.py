"""Export helpers — CSV / YAML / Excel (ported from gui/views/viewer.py)."""

from __future__ import annotations

import csv
import io
import re
import sqlite3
from pathlib import Path


def _yaml_scalar(v):
    s = str(v) if v is not None else ""
    if not s:
        return '""'
    if (re.search(r'[:{}\[\]#&*!,|>\'"@`]', s)
            or s[0] in "-?~"
            or s.lower() in ("true", "false", "null", "yes", "no", "on", "off")
            or re.match(r'^-?\d', s)):
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return s


def _rows_to_yaml_fallback(headers: list[str], rows: list[list[str]]) -> str:
    lines = []
    for row in [headers, *rows]:
        items = ", ".join(_yaml_scalar(v) for v in row)
        lines.append(f"- [{items}]")
    return "\n".join(lines) + "\n"


def make_csv(headers: list[str], rows: list[list[str]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter="|", lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue()


def make_yaml(headers: list[str], rows: list[list[str]]) -> str:
    try:
        import yaml

        class _FlowList(list):
            pass

        def _flow_representer(dumper, data):
            return dumper.represent_sequence(
                "tag:yaml.org,2002:seq", data, flow_style=True)

        yaml.add_representer(_FlowList, _flow_representer)
        data = [_FlowList(headers)] + [_FlowList(row) for row in rows]
        return yaml.dump(data, default_flow_style=False, allow_unicode=True)
    except ImportError:
        return _rows_to_yaml_fallback(headers, rows)


def make_excel(
    headers: list[str],
    rows: list[list[str]],
    categories: list[dict],
    sheet_title: str = "Attack Paths",
) -> bytes:
    """Build an openpyxl workbook with classification-based row highlighting."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cls_fills: dict[str, PatternFill] = {}
    for cat in categories:
        hex_color = cat["color"].lstrip("#").upper()
        if len(hex_color) == 3:
            hex_color = "".join(c * 2 for c in hex_color)
        cls_fills[cat["name"]] = PatternFill("solid", fgColor=hex_color)

    cls_col_idx: int | None = None
    for ci, h in enumerate(headers, 1):
        if h.lower() == "classification":
            cls_col_idx = ci
            break

    ws.append(headers)
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for dr in rows:
        ws.append([v if v is not None else "" for v in dr])

    if cls_col_idx is not None:
        for ri in range(2, ws.max_row + 1):
            cls_val = ws.cell(ri, cls_col_idx).value
            if cls_val and cls_val in cls_fills:
                row_fill = cls_fills[cls_val]
                for ci in range(1, len(headers) + 1):
                    ws.cell(ri, ci).fill = row_fill

    for ci, hdr in enumerate(headers, 1):
        max_len = len(str(hdr))
        for ri in range(2, ws.max_row + 1):
            v = ws.cell(ri, ci).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fetch_full_policies_with_annotations(db_path: Path) -> tuple[list[str], list[list]]:
    headers = [
        "Policy Name", "Source Zone", "Dest Zone", "Source",
        "Destination", "Application", "Service", "Action",
        "Classification", "Notes",
    ]
    conn = sqlite3.connect(db_path)
    rows = conn.execute(
        "SELECT p.policyname, p.source_zone, p.destn_zone,"
        "       p.source, p.destination, p.application,"
        "       p.service, p.action,"
        "       a.classification, a.notes"
        " FROM  t_policies p"
        " LEFT JOIN t_annotations a"
        "        ON a.policy_name = p.policyname"
        " ORDER BY p.rowid"
    ).fetchall()
    conn.close()
    return headers, [list(r) for r in rows]
